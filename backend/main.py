from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
import json
import re
import io
import uuid
from datetime import datetime, timedelta
from pathlib import Path
import os

app = FastAPI(title="Impact Analysis - Demo")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)

_default_data = Path(__file__).parent.parent / "data"
DATA_DIR = Path(os.getenv("DATA_DIR", str(_default_data)))
DATA_DIR.mkdir(parents=True, exist_ok=True)
LOCK_TIMEOUT_MINUTES = 30


# ── Models ────────────────────────────────────────────────────────────────────

class FailureMode(BaseModel):
    id: Optional[str] = None
    function_feature: str = ""
    failure_mode: str = ""
    dependency_severity: str = ""
    effect_on_service: str = ""
    effect_on_guest: str = ""
    effect_on_worker: str = ""
    sev: Optional[int] = None
    occ: Optional[int] = None
    det: Optional[int] = None
    runbook_links: str = ""


class LockRequest(BaseModel):
    user_name: str
    session_id: str


class CreateCIRequest(BaseModel):
    ci_id: str


class SaveRequest(BaseModel):
    session_id: str
    failure_modes: List[FailureMode]


# ── Helpers ───────────────────────────────────────────────────────────────────

def safe_filename(ci_id: str) -> Path:
    safe = re.sub(r"[^\w\-]", "_", ci_id)
    return DATA_DIR / f"{safe}.json"


def find_ci_file(ci_id: str) -> Optional[Path]:
    direct = safe_filename(ci_id)
    if direct.exists():
        try:
            with open(direct) as f:
                if json.load(f).get("ci_id") == ci_id:
                    return direct
        except Exception:
            pass
    for path in DATA_DIR.glob("*.json"):
        try:
            with open(path) as f:
                if json.load(f).get("ci_id") == ci_id:
                    return path
        except Exception:
            pass
    return None


def load_ci(ci_id: str) -> dict:
    path = find_ci_file(ci_id)
    if not path:
        raise HTTPException(status_code=404, detail=f"CI '{ci_id}' not found")
    with open(path) as f:
        return json.load(f)


def save_ci(data: dict) -> None:
    with open(safe_filename(data["ci_id"]), "w") as f:
        json.dump(data, f, indent=2, default=str)


def is_expired(lock: Optional[dict]) -> bool:
    if not lock:
        return True
    return datetime.utcnow() > datetime.fromisoformat(lock["expires_at"])


def make_lock(user_name: str, session_id: str) -> dict:
    now = datetime.utcnow()
    return {
        "user_name": user_name,
        "session_id": session_id,
        "locked_at": now.isoformat(),
        "expires_at": (now + timedelta(minutes=LOCK_TIMEOUT_MINUTES)).isoformat(),
    }


def calc_rpn(fm: dict) -> int:
    return (fm.get("sev") or 0) * (fm.get("occ") or 0) * (fm.get("det") or 0)


def lock_summary(lock: Optional[dict]) -> Optional[dict]:
    if not lock or is_expired(lock):
        return None
    return {"user_name": lock["user_name"], "expires_at": lock["expires_at"]}


# ── Routes ────────────────────────────────────────────────────────────────────

@app.get("/api/cis")
def list_cis():
    result = []
    for path in DATA_DIR.glob("*.json"):
        try:
            with open(path) as f:
                data = json.load(f)
            result.append({
                "ci_id": data["ci_id"],
                "created_at": data.get("created_at"),
                "updated_at": data.get("updated_at"),
                "failure_mode_count": len(data.get("failure_modes", [])),
                "lock": lock_summary(data.get("lock")),
            })
        except Exception:
            pass
    return sorted(result, key=lambda x: x["ci_id"].lower())


@app.post("/api/cis", status_code=201)
def create_ci(req: CreateCIRequest):
    ci_id = req.ci_id.strip()
    if not ci_id:
        raise HTTPException(400, "CI ID cannot be empty")
    if not re.match(r"^[\w\-\.]+$", ci_id):
        raise HTTPException(400, "CI ID may only contain letters, numbers, hyphens, underscores, and dots")
    if find_ci_file(ci_id):
        raise HTTPException(409, f"CI '{ci_id}' already exists")
    now = datetime.utcnow().isoformat()
    data = {
        "ci_id": ci_id,
        "created_at": now,
        "updated_at": now,
        "lock": None,
        "failure_modes": [],
    }
    save_ci(data)
    return data


@app.get("/api/cis/{ci_id}")
def get_ci(ci_id: str):
    data = load_ci(ci_id)
    if data.get("lock") and is_expired(data["lock"]):
        data["lock"] = None
        save_ci(data)
    return data


@app.post("/api/cis/{ci_id}/lock")
def acquire_lock(ci_id: str, req: LockRequest):
    data = load_ci(ci_id)
    lock = data.get("lock")
    if lock and not is_expired(lock) and lock["session_id"] != req.session_id:
        raise HTTPException(
            status_code=423,
            detail={
                "message": f"Locked by {lock['user_name']}",
                "locked_by": lock["user_name"],
                "expires_at": lock["expires_at"],
            },
        )
    data["lock"] = make_lock(req.user_name, req.session_id)
    save_ci(data)
    return {"status": "locked", "lock": data["lock"]}


@app.post("/api/cis/{ci_id}/heartbeat")
def heartbeat(ci_id: str, req: LockRequest):
    data = load_ci(ci_id)
    lock = data.get("lock")
    if not lock or lock["session_id"] != req.session_id:
        raise HTTPException(403, "Session does not hold the lock")
    data["lock"] = make_lock(req.user_name, req.session_id)
    save_ci(data)
    return {"status": "renewed"}


@app.post("/api/cis/{ci_id}/unlock")
def release_lock(ci_id: str, req: LockRequest):
    data = load_ci(ci_id)
    lock = data.get("lock")
    if lock and lock["session_id"] == req.session_id:
        data["lock"] = None
        save_ci(data)
    return {"status": "unlocked"}


@app.put("/api/cis/{ci_id}/failure-modes")
def save_failure_modes(ci_id: str, req: SaveRequest):
    data = load_ci(ci_id)
    lock = data.get("lock")
    if not lock or lock["session_id"] != req.session_id or is_expired(lock):
        raise HTTPException(403, "You do not hold the edit lock for this CI")
    fms = []
    for i, fm in enumerate(req.failure_modes):
        d = fm.model_dump()
        d["id"] = d.get("id") or str(uuid.uuid4())
        d["item_number"] = i + 1
        d["rpn"] = calc_rpn(d)
        fms.append(d)
    data["failure_modes"] = fms
    data["updated_at"] = datetime.utcnow().isoformat()
    save_ci(data)
    return data


@app.get("/api/cis/{ci_id}/export")
def export_ci(ci_id: str):
    data = load_ci(ci_id)
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r'[\\/*?\[\]:]', '_', ci_id)[:31]

    headers = [
        "Item #",
        "Function / Feature of my service",
        "Failure Mode of Upstream Dependency",
        "Upstream Dependency Severity",
        "Effect on my service",
        "Effect on Guest",
        "Effect on Worker",
        "Severity (SEV)",
        "Occurrence (OCC)",
        "Detection (DET)",
        "RPN (SEV × OCC × DET)",
        "Related Runbook(s) Link",
    ]
    widths = [8, 35, 35, 25, 35, 25, 25, 14, 14, 14, 20, 40]

    green_fill = PatternFill("solid", fgColor="217346")
    white_bold = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_wrap = Alignment(vertical="top", wrap_text=True)

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = green_fill
        cell.font = white_bold
        cell.alignment = center_wrap
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 40

    for fm in data.get("failure_modes", []):
        ws.append([
            fm.get("item_number", ""),
            fm.get("function_feature", ""),
            fm.get("failure_mode", ""),
            fm.get("dependency_severity", ""),
            fm.get("effect_on_service", ""),
            fm.get("effect_on_guest", ""),
            fm.get("effect_on_worker", ""),
            fm.get("sev") or "",
            fm.get("occ") or "",
            fm.get("det") or "",
            fm.get("rpn") or "",
            fm.get("runbook_links", ""),
        ])
        for c in range(1, 13):
            ws.cell(ws.max_row, c).alignment = top_wrap

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = re.sub(r"[^\w\-]", "_", ci_id)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}_impact_analysis.xlsx"'},
    )


# ── Static files (must be last) ───────────────────────────────────────────────

frontend_dist = Path(__file__).parent.parent / "frontend" / "dist"
if frontend_dist.exists():
    app.mount("/", StaticFiles(directory=str(frontend_dist), html=True), name="static")
